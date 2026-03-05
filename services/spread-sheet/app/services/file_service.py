from io import BytesIO
import pandas as pd
import openpyxl

class FileService:
    def __init__(self):
        self.wb_formulas = None
        self.wb_values = None
    
    def read_excel_file(self, file_content, sheet_names=None, all_sheets=False):
        """
        Read Excel file content with multiple sheets handling
        
        Parameters:
        file_content (bytes): The Excel file content as bytes
        sheet_names (list): List of specific sheet names to read. If None and all_sheets=False, reads first sheet
        all_sheets (bool): If True, reads all sheets regardless of sheet_names parameter
        
        Returns:
        dict: Dictionary of DataFrames with sheet names as keys, or single DataFrame if only one sheet
        """
        excel_file = BytesIO(file_content)
        
        # Load workbook
        self.wb_formulas = openpyxl.load_workbook(excel_file, data_only=False)
        self.wb_values = openpyxl.load_workbook(excel_file, data_only=True)
        
        # Get all available sheet names
        available_sheets = pd.ExcelFile(excel_file).sheet_names
        
        if all_sheets:
            # Read all sheets
            dfs = pd.read_excel(excel_file, sheet_name=None)
        elif sheet_names:
            # Validate requested sheets exist
            invalid_sheets = [sheet for sheet in sheet_names if sheet not in available_sheets]
            if invalid_sheets:
                raise ValueError(f"Sheets not found: {invalid_sheets}")
            
            # Read specific sheets
            dfs = pd.read_excel(excel_file, sheet_name=sheet_names)
        else:
            # Read only the first sheet
            dfs = pd.read_excel(excel_file)
            
        # Clean DataFrames
        # dfs = self.clean_xlsx(dfs)
            
        # Ensure dfs is a dictionary for consistent processing
        if not isinstance(dfs, dict):
            dfs = {available_sheets[0]: dfs}
        
        return dfs
    
    def clean_xlsx(self, dfs):
        """
        Clean Excel dataframes by removing empty rows/columns and handling multiple empty header rows.
        
        Parameters:
            dfs (dict): Dictionary of DataFrames with sheet names as keys
        Returns:
            dict: Dictionary of cleaned DataFrames
        """
        cleaned_dfs = {}
        
        for sheet_name, df in dfs.items():
            # Remove empty columns only
            df_cleaned = df.dropna(axis=1, how='all')
            
            if df_cleaned.empty:
                cleaned_dfs[sheet_name] = df_cleaned
                continue
                
            # Check if we already have valid headers
            current_headers = list(df_cleaned.columns)
            if all(str(header).startswith('Unnamed:') or str(header).isdigit() 
                for header in current_headers):
                # Only look for headers in data if current headers are invalid
                first_valid_row = None
                for idx, row in df_cleaned.iterrows():
                    if not row.isnull().all():
                        if any(isinstance(val, str) for val in row if pd.notna(val)):
                            first_valid_row = idx
                            break
                
                if first_valid_row is not None:
                    headers = df_cleaned.iloc[first_valid_row]
                    df_cleaned = df_cleaned.drop(first_valid_row).reset_index(drop=True)
                    df_cleaned = df_cleaned[first_valid_row:].reset_index(drop=True)
                    df_cleaned.columns = headers
            
            # Final cleanup of empty rows
            df_cleaned = df_cleaned.dropna(axis=0, how='all').reset_index(drop=True)
            cleaned_dfs[sheet_name] = df_cleaned
        
        return cleaned_dfs
    
    def read_csv_file(self, file_content):
        """
        Read CSV file content into a DataFrame
        
        Parameters:
        file_content (bytes): The CSV file content as bytes
        
        Returns:
        dict: Dictionary of DataFrames with 'csv_default_sheet_name' as key
        """
        csv_file = BytesIO(file_content)
        
        # Use pandas to read the CSV content
        df = pd.read_csv(csv_file)
        
        return {"csv_default_sheet_name": df}
    
    def get_cell_formular(self, cell_address, sheet_name=None):
        """
        Get the formula and value from a specific cell across all or specific sheets
        
        Parameters:
        cell_address (str): The cell address (e.g., 'A1', 'B2')
        sheet_name (str, optional): Specific sheet name to check. If None, checks all sheets
        
        Returns:
        dict: Dictionary with sheet names as keys and formula information as values
            Format: {
                'sheet_name': {
                    'formula': 'formula_string_or_None',
                    'value': 'cell_value',
                    'error': 'error_message_if_any'
                }
            }
        """
        result = {}
        sheets_to_check = [sheet_name] if sheet_name else self.wb_formulas.sheetnames
        
        for sheet_name in sheets_to_check:
            if sheet_name not in self.wb_formulas.sheetnames:
                result[sheet_name] = {
                        'formula': None,
                        'value': None,
                        'error': f'Sheet "{sheet_name}" not found'
                    }
                continue
            
            sheet_formulas = self.wb_formulas[sheet_name]
            sheet_values = self.wb_values[sheet_name]
            
            try:
                # Get cell for formula
                cell_formula = sheet_formulas[cell_address]
                # Get cell for value
                cell_value = sheet_values[cell_address]
                
                # Check if cell has a formula by checking if it starts with '='
                formula_str = cell_formula.value
                formula = formula_str if isinstance(formula_str, str) and formula_str.startswith('=') else None
                
                result[sheet_name] = {
                    'formula': formula,
                    'value': cell_value.value,
                    'error': None
                }
            except Exception as e:
                result[sheet_name] = {
                    'formula': None,
                    'value': None,
                    'error': f'Invalid cell address: {str(e)}'
                }
            
        return result
            
    def extract_file_content_bytes(self, dfs):
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name, df in dfs.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
        output.seek(0)
        return output.read()